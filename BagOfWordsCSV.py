################################################## IMPORTACIONES #######################################################
import json
import openpyxl
# LIBRERIA PARA REALIZAR LA MATRIZ DE BAG OF WORDS
from keras.preprocessing.text import Tokenizer

# LIBRERIA PARA REALIZAR NUBES DE PALABRAS
import stylecloud as sc

rango = "7000-8040"

############################################## GENERAR BAG OF WORDS ####################################################

# RECUPERAR EL ARREGLO DE WORDS GENERAL
array_words = []
nombre_leyes = []

# RECUPERAMOS LA URL DEL JSON
url_creator = "./PREPROCESADO-LEYES/JSON/"+rango+".json"

# ABRIMOS EL ARCHIVO JSON
with open(url_creator, encoding="utf-8") as file:
    data = json.load(file)

    # RECORREMOS EL JSON Y RECUPERAMOS EL ARREGLO DE WORDS Y EL IDENTIFICADOR DE LA LEY
    for aux in data['leyes']:
        array_words.append( aux['words'] )
        nombre_leyes.append( aux['ley'] )

# REALIZAMOS EL FIT DE LAS PALABRAS DEL ARREGLO PREPROCESADO
model = Tokenizer()
model.fit_on_texts(array_words)

# ARREGLO DE PALABRAS PREPROCESADAS
keys = list(model.word_index.keys())

# REALIZAMOS LA MATRIZ DE BAG OF WORDS
rep = model.texts_to_matrix(array_words, mode='count')

################################################ CONSTRUIMOS EXCEL #####################################################

# HABILITAMOS LA LIBRERIA QUE NOS PERMITE TRABAJAR CON UN ARCHIVO EXCEL
wb = openpyxl.Workbook()
hoja = wb.active

# GENERAMOS LAS FILAS USANDO LA MATRIZ DE BAG OF WORDS
filas = [ ]
for i in range (len(nombre_leyes)):
    aux = list(rep[i])
    del aux[0]
    fila = [nombre_leyes[i]] + aux
    filas.append(fila)

# CREAMOS LA FILA CON LOS ENCABEZADOS (ARRAY_WORDS)
array = [''] + keys
hoja.append(array)




# AGREGAMOS LAS FILAS A NUESTRA HOJA DE EXCEL
for producto in filas:
    # producto es una tupla con los valores de un producto
    hoja.append(producto)
    print( producto[0] )


# REDIMENSIONAMOS EL ARCHIVO EXCEL
for column_cells in hoja.columns:
    new_column_letter = (openpyxl.utils.get_column_letter(column_cells[0].column))
    hoja.column_dimensions[new_column_letter].width = 12


# GUARDAMOS EL ARCHIVO EXCEL CREADO
url_save = './PREPROCESADO-LEYES/EXCEL/' + rango + ".xlsx"
wb.save(url_save)



################################################ NUBE DE PALABRAS ######################################################

'''
# GENERAMOS EL TEXTO, USANDO LOS ARRAY_WORDS
texto = ""
for j in array_words:
    for i in range (len(j)):
        texto += j[i] + " "
'''

# DICCIONARIO PARA LA NUBE DE PALABRAS
dic_ponderado = {}

lista = list(model.word_docs.items())

for i in range(len(lista)):
    dic_ponderado[str(lista[i][0])] = lista[i][1]

# ORDENAMOS DICCIONARIO DE PALABRAS PREPROCESADAS PARA PASARLO A UN EXCEL
sorted_words = sorted(dic_ponderado.items(), key=lambda x: x[1], reverse=True)
wb1 = openpyxl.Workbook()
hoja1 = wb1.active
for i in range(len(sorted_words)):
    fila = [ sorted_words[i][0], sorted_words[i][1]]
    hoja1.append(fila)
url_save = './PREPROCESADO-LEYES/EXCEL-TF/' + rango + ".xlsx"
wb1.save(url_save)


# NUMERO DE PALABRAS
print(len(keys))

# CREAMOS LA NUBE DE PALABRAS Y LA GUARDAMOS
url_save_png = './PREPROCESADO-LEYES/NUBE DE PALABRAS-TF/' + rango + ".png"

sc.gen_stylecloud(
    text= dic_ponderado,
    colors=['#ecf0f1', '#3498db', '#e74c3c'],
    icon_name='fas fa-cloud',
    background_color='#1A1A1A',
    output_name = url_save_png,
    size = 2048,
    max_words = 300,
    max_font_size = 450
)